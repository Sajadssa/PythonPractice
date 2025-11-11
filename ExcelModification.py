import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# داده‌های WBS
tasks = [
    ['0.0', 'Setup', 'Project Initialization', 'راه‌اندازی اولیه پروژه', '-', 2, 'Critical', 'Pending'],
    ['0.1', 'Setup', 'Install Docker Desktop', 'نصب و کانفیگ Docker', '0.0', 1, 'Critical', 'Pending'],
    ['0.2', 'Setup', 'Clean Old Files', 'حذف فایل‌های قدیمی', '0.1', 0.5, 'High', 'Pending'],
    
    ['1.0', 'Infrastructure', 'Create Base Structure', 'ایجاد ساختار پوشه‌بندی', '0.2', 1, 'Critical', 'Pending'],
    ['1.1', 'Infrastructure', 'Docker Compose - Base', 'ایجاد docker-compose.yml پایه', '1.0', 2, 'Critical', 'Pending'],
    ['1.2', 'Infrastructure', 'Environment Variables', 'ساخت .env و .env.example', '1.1', 1, 'High', 'Pending'],
    ['1.3', 'Infrastructure', 'PostgreSQL Service', 'کانفیگ PostgreSQL در Docker', '1.2', 2, 'Critical', 'Pending'],
    ['1.4', 'Infrastructure', 'Redis Service', 'کانفیگ Redis در Docker', '1.3', 1, 'Medium', 'Pending'],
    ['1.5', 'Infrastructure', 'Network Configuration', 'راه‌اندازی Docker networks', '1.4', 1, 'High', 'Pending'],
    ['1.6', 'Infrastructure', 'Volume Configuration', 'کانفیگ persistent volumes', '1.5', 1, 'High', 'Pending'],
    ['1.7', 'Infrastructure', 'Test Infrastructure', 'تست اتصالات و سرویس‌ها', '1.6', 1, 'High', 'Pending'],
    
    ['2.0', 'Backend', 'Backend Dockerfile', 'ساخت Dockerfile برای FastAPI', '1.7', 2, 'Critical', 'Pending'],
    ['2.1', 'Backend', 'Requirements File', 'لیست dependencies پایتون', '2.0', 1, 'Critical', 'Pending'],
    ['2.2', 'Backend', 'FastAPI Structure', 'ساختار اولیه FastAPI app', '2.1', 3, 'Critical', 'Pending'],
    ['2.3', 'Backend', 'Database Configuration', 'کانفیگ SQLAlchemy + Alembic', '2.2', 2, 'Critical', 'Pending'],
    ['2.4', 'Backend', 'Base Models', 'ساخت Base و Mixins برای Models', '2.3', 2, 'High', 'Pending'],
    ['2.5', 'Backend', 'Authentication System', 'JWT + OAuth2', '2.4', 4, 'Critical', 'Pending'],
    ['2.6', 'Backend', 'User Management', 'CRUD کاربران', '2.5', 3, 'High', 'Pending'],
    ['2.7', 'Backend', 'RFI Models', 'Models مربوط به RFI', '2.6', 3, 'Critical', 'Pending'],
    ['2.8', 'Backend', 'RFI CRUD Operations', 'ایجاد/ویرایش/حذف RFI', '2.7', 4, 'Critical', 'Pending'],
    ['2.9', 'Backend', 'Response Models', 'Models پاسخ‌ها', '2.8', 2, 'High', 'Pending'],
    ['2.10', 'Backend', 'Response CRUD', 'مدیریت پاسخ‌ها', '2.9', 3, 'High', 'Pending'],
    ['2.11', 'Backend', 'File Upload Service', 'آپلود فایل‌های ضمیمه', '2.10', 3, 'Medium', 'Pending'],
    ['2.12', 'Backend', 'Search & Filter', 'جستجو و فیلتر پیشرفته', '2.11', 3, 'High', 'Pending'],
    ['2.13', 'Backend', 'Pagination Service', 'صفحه‌بندی داده‌ها', '2.12', 2, 'High', 'Pending'],
    ['2.14', 'Backend', 'Validation Layer', 'اعتبارسنجی داده‌ها', '2.13', 2, 'Medium', 'Pending'],
    ['2.15', 'Backend', 'Error Handling', 'مدیریت خطاها', '2.14', 2, 'High', 'Pending'],
    ['2.16', 'Backend', 'Logging System', 'سیستم لاگ', '2.15', 2, 'Medium', 'Pending'],
    ['2.17', 'Backend', 'API Documentation', 'مستندسازی Swagger', '2.16', 2, 'Low', 'Pending'],
    ['2.18', 'Backend', 'Unit Tests', 'تست‌های واحد', '2.17', 4, 'High', 'Pending'],
    ['2.19', 'Backend', 'Backend Integration', 'تست یکپارچه‌سازی', '2.18', 2, 'High', 'Pending'],
    
    ['3.0', 'Frontend', 'Frontend Dockerfile', 'ساخت Dockerfile برای Next.js', '2.19', 2, 'Critical', 'Pending'],
    ['3.1', 'Frontend', 'Next.js Base Setup', 'راه‌اندازی اولیه Next.js', '3.0', 2, 'Critical', 'Pending'],
    ['3.2', 'Frontend', 'Tailwind Configuration', 'کانفیگ Tailwind CSS', '3.1', 1, 'High', 'Pending'],
    ['3.3', 'Frontend', 'TypeScript Setup', 'تنظیمات TypeScript', '3.2', 1, 'High', 'Pending'],
    ['3.4', 'Frontend', 'Folder Structure', 'ساختار پوشه‌های frontend', '3.3', 1, 'High', 'Pending'],
    ['3.5', 'Frontend', 'API Service Layer', 'سرویس HTTP با Axios', '3.4', 2, 'Critical', 'Pending'],
    ['3.6', 'Frontend', 'Auth Context', 'Context مدیریت Authentication', '3.5', 3, 'Critical', 'Pending'],
    ['3.7', 'Frontend', 'RFI Context', 'Context مدیریت RFI', '3.6', 3, 'Critical', 'Pending'],
    ['3.8', 'Frontend', 'UI Components - Base', 'کامپوننت‌های پایه', '3.7', 4, 'High', 'Pending'],
    ['3.9', 'Frontend', 'Layout Components', 'Header/Sidebar/Footer', '3.8', 3, 'High', 'Pending'],
    ['3.10', 'Frontend', 'Login Page', 'صفحه ورود', '3.9', 3, 'Critical', 'Pending'],
    ['3.11', 'Frontend', 'Dashboard Page', 'داشبورد اصلی', '3.10', 4, 'Critical', 'Pending'],
    ['3.12', 'Frontend', 'RFI List Page', 'لیست RFI با جدول', '3.11', 4, 'Critical', 'Pending'],
    ['3.13', 'Frontend', 'RFI Form', 'فرم ایجاد/ویرایش RFI', '3.12', 4, 'Critical', 'Pending'],
    ['3.14', 'Frontend', 'RFI Detail Page', 'جزئیات RFI', '3.13', 3, 'High', 'Pending'],
    ['3.15', 'Frontend', 'Response List', 'لیست پاسخ‌ها', '3.14', 3, 'High', 'Pending'],
    ['3.16', 'Frontend', 'Response Form', 'فرم ثبت پاسخ', '3.15', 3, 'High', 'Pending'],
    ['3.17', 'Frontend', 'Search Component', 'جستجوی پیشرفته', '3.16', 3, 'Medium', 'Pending'],
    ['3.18', 'Frontend', 'Filter Component', 'فیلترهای چندگانه', '3.17', 3, 'Medium', 'Pending'],
    ['3.19', 'Frontend', 'Pagination Component', 'کامپوننت صفحه‌بندی', '3.18', 2, 'High', 'Pending'],
    ['3.20', 'Frontend', 'File Upload Component', 'آپلود فایل', '3.19', 3, 'Medium', 'Pending'],
    ['3.21', 'Frontend', 'Notification System', 'سیستم اعلان‌ها', '3.20', 2, 'Medium', 'Pending'],
    ['3.22', 'Frontend', 'Loading States', 'حالت‌های بارگذاری', '3.21', 2, 'Medium', 'Pending'],
    ['3.23', 'Frontend', 'Error Boundaries', 'مدیریت خطاها در UI', '3.22', 2, 'High', 'Pending'],
    ['3.24', 'Frontend', 'Responsive Design', 'بهینه‌سازی موبایل', '3.23', 3, 'High', 'Pending'],
    ['3.25', 'Frontend', 'Frontend Tests', 'تست‌های کامپوننت', '3.24', 3, 'Medium', 'Pending'],
    
    ['4.0', 'Integration', 'API Integration', 'اتصال Frontend به Backend', '3.25', 3, 'Critical', 'Pending'],
    ['4.1', 'Integration', 'Authentication Flow', 'تست فرآیند احراز هویت', '4.0', 2, 'Critical', 'Pending'],
    ['4.2', 'Integration', 'CRUD Operations Test', 'تست عملیات CRUD', '4.1', 3, 'Critical', 'Pending'],
    ['4.3', 'Integration', 'File Upload Test', 'تست آپلود فایل', '4.2', 2, 'Medium', 'Pending'],
    ['4.4', 'Integration', 'Search & Filter Test', 'تست جستجو و فیلتر', '4.3', 2, 'High', 'Pending'],
    ['4.5', 'Integration', 'Error Handling Test', 'تست مدیریت خطا', '4.4', 2, 'High', 'Pending'],
    
    ['5.0', 'DevOps', 'Nginx Configuration', 'کانفیگ Nginx', '4.5', 2, 'High', 'Pending'],
    ['5.1', 'DevOps', 'Multi-stage Builds', 'بهینه‌سازی Dockerfiles', '5.0', 2, 'Medium', 'Pending'],
    ['5.2', 'DevOps', 'Docker Compose - Production', 'کانفیگ production', '5.1', 2, 'High', 'Pending'],
    ['5.3', 'DevOps', 'Health Checks', 'تست سلامت سرویس‌ها', '5.2', 1, 'Medium', 'Pending'],
    ['5.4', 'DevOps', 'Backup Strategy', 'استراتژی پشتیبان‌گیری', '5.3', 2, 'Medium', 'Pending'],
    ['5.5', 'DevOps', 'Monitoring Setup', 'راه‌اندازی مانیتورینگ', '5.4', 3, 'Low', 'Pending'],
    ['5.6', 'DevOps', 'CI/CD Pipeline', 'خط لوله CI/CD', '5.5', 4, 'Medium', 'Pending'],
    
    ['6.0', 'Documentation', 'Technical Documentation', 'مستندات فنی', '5.6', 3, 'High', 'Pending'],
    ['6.1', 'Documentation', 'API Documentation', 'مستندات API', '6.0', 2, 'High', 'Pending'],
    ['6.2', 'Documentation', 'User Guide', 'راهنمای کاربر', '6.1', 3, 'Medium', 'Pending'],
    ['6.3', 'Documentation', 'Deployment Guide', 'راهنمای استقرار', '6.2', 2, 'High', 'Pending'],
    ['6.4', 'Documentation', 'README Files', 'فایل‌های README', '6.3', 1, 'High', 'Pending'],
    
    ['7.0', 'Testing', 'Integration Testing', 'تست یکپارچه‌سازی کامل', '6.4', 4, 'Critical', 'Pending'],
    ['7.1', 'Testing', 'Performance Testing', 'تست عملکرد', '7.0', 3, 'High', 'Pending'],
    ['7.2', 'Testing', 'Security Testing', 'تست امنیتی', '7.1', 3, 'High', 'Pending'],
    ['7.3', 'Testing', 'User Acceptance Testing', 'تست پذیرش کاربر', '7.2', 4, 'High', 'Pending'],
    
    ['8.0', 'Deployment', 'Production Deployment', 'استقرار نهایی', '7.3', 4, 'Critical', 'Pending'],
    ['8.1', 'Deployment', 'Post-deployment Testing', 'تست بعد از استقرار', '8.0', 2, 'Critical', 'Pending'],
    ['8.2', 'Deployment', 'Handover', 'تحویل پروژه', '8.1', 2, 'High', 'Pending']
]

df = pd.DataFrame(tasks, columns=['Task_ID', 'Phase', 'Task_Name', 'Description', 'Dependencies', 'Duration_Hours', 'Priority', 'Status'])

df['Start_Date'] = ''
df['End_Date'] = ''
df['Assigned_To'] = ''
df['Notes'] = ''
df['Completion_%'] = 0

output_file = 'IDMS_WRFM_WBS.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='WBS', index=False)
    
    workbook = writer.book
    worksheet = writer.sheets['WBS']
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    priority_colors = {
        'Critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
        'High': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
        'Medium': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
        'Low': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    }
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        priority_cell = row[6]
        if priority_cell.value in priority_colors:
            priority_cell.fill = priority_colors[priority_cell.value]
    
    column_widths = {'A': 10, 'B': 15, 'C': 30, 'D': 35, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 20, 'L': 40, 'M': 12}
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    worksheet.freeze_panes = 'A2'

print(f"✅ فایل {output_file} ساخته شد")
print(f"📊 تعداد تسک‌ها: {len(df)}")
print(f"⏱️ کل زمان: {df['Duration_Hours'].sum()} ساعت = {df['Duration_Hours'].sum()/8:.1f} روز")

# تحلیل WBS
print("\n📊 تحلیل پروژه:")
print(f"├─ تعداد کل تسک‌ها: {len(df)}")
print(f"├─ تخمین زمان کل: {df['Duration_Hours'].sum()} ساعت")
print(f"├─ معادل روز کاری: {df['Duration_Hours'].sum()/8:.1f} روز")
print(f"└─ معادل هفته کاری: {df['Duration_Hours'].sum()/40:.1f} هفته")

print("\n🎯 تسک‌های Critical:")
critical_tasks = df[df['Priority'] == 'Critical']
print(f"├─ تعداد: {len(critical_tasks)}")
print(f"└─ زمان: {critical_tasks['Duration_Hours'].sum()} ساعت")

print("\n📦 توزیع فازها:")
phase_summary = df.groupby('Phase').agg({
    'Task_ID': 'count',
    'Duration_Hours': 'sum'
}).rename(columns={'Task_ID': 'Tasks', 'Duration_Hours': 'Hours'})
print(phase_summary)
