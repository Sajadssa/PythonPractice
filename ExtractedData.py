"""
Spare Parts List Extractor
This script extracts parts data from images and exports to Excel/CSV
Based on the standard form with columns as per image specifications
"""

import pytesseract
from PIL import Image
import pandas as pd
import re
import cv2
import numpy as np
import os

# If Tesseract is installed in a different location, specify the path
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def preprocess_image(image_path):
    """
    Preprocess image for better OCR results
    
    Args:
        image_path: Path to the input image file
        
    Returns:
        Preprocessed image array
    """
    if not os.path.exists(image_path):
        print(f"Warning: Image file '{image_path}' not found!")
        return None
        
    img = cv2.imread(image_path)
    
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Apply threshold for better quality
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
    
    # Denoise
    denoised = cv2.fastNlMeansDenoising(thresh, None, 10, 7, 21)
    
    return denoised

def extract_table_data_ocr(image_path):
    """
    Extract table data from image using OCR
    
    Args:
        image_path: Path to the input image file
        
    Returns:
        List of dictionaries containing extracted data
    """
    
    # Preprocess image
    processed_img = preprocess_image(image_path)
    
    if processed_img is None:
        return []
    
    # Extract text using OCR
    text = pytesseract.image_to_string(processed_img, lang='eng', config='--psm 6')
    
    # Process text and extract information
    lines = text.split('\n')
    
    data = []
    
    for line in lines:
        # Filter empty lines
        if not line.strip():
            continue
            
        # Search for pattern with PCS
        if 'PCS' in line or 'pcs' in line.lower():
            
            try:
                # Extract basic information
                unit = 'PCS.'
                
                # Find numbers in line
                numbers = re.findall(r'\d+', line)
                
                if len(numbers) >= 2:
                    no_of_units = int(numbers[0])
                    qty = int(numbers[1])
                    
                    # Extract part description
                    desc_match = re.search(r'PCS\.\s*\d+\s+(.+?)(?:SIEC-|$)', line)
                    description = desc_match.group(1).strip() if desc_match else ''
                    
                    # Extract drawing reference number
                    ref_match = re.search(r'(SIEC-[A-Z0-9-]+)', line)
                    drawing_ref = ref_match.group(1) if ref_match else 'SIEC-DGNRAC-ENEL-DIWI-0001'
                    
                    # Add to data with column names from the form
                    data.append({
                        'UNIT (Kg, No., pair, set)': unit,
                        'Total number of identical parts installed': no_of_units,
                        'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': description,
                        'Drawing/Ref No': drawing_ref,
                        'Material (See Note 4 Above)': 'Multimaterial',
                        'Recommended by manufacturer': '',
                    })
                    
            except (ValueError, IndexError):
                continue
    
    return data

def get_predefined_data():
    """
    Returns predefined data extracted from the document
    Column names based on the standard form
    
    Returns:
        List of dictionaries with parts data
    """
    return [
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 40,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'FUSE, 10x38 mm, 2A',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 11
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 4,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'FUSE, 10x38 mm, 4A',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 4,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Power Fuse, 630A',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 2,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'MCB - 6A, 1P, for heater & lighting',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 31,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'MCB - 10A, 2P, for distribution',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 5
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 36,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'MCB - 16A, 2P, for distribution',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 5
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 5,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'MCB - 20A, 2P, for distribution',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 5,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'MCB - 32A, 2P, for distribution',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 12,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Rec. Thyristor (SCR), SKKT323',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 4
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 4,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'STS Thyristor (SCR), SKKT323',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 2,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Block Diode Module',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 2,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'IGBT (Transistor), CM1400',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 4,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Rectifier Thyristor Driver PCB (control cards)',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 2,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'IGBT Driver PCB (control cards)',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 2,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'STS Thyristor Driver PCB (control cards)',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 1,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Isolator switch (make before break), 400A',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 1
        },
        {
            'UNIT (Kg, No., pair, set)': 'PCS.',
            'Total number of identical parts installed': 352,
            'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)': 'Batteries (SBM625, 176 Cells) (2 Banks)',
            'Drawing/Ref No': 'SIEC-DGNRAC-ENEL-DIWI-0001',
            'Material (See Note 4 Above)': 'Multimaterial',
            'Recommended by manufacturer': 18
        }
    ]

def save_to_excel(df, filename='parts_list_extracted.xlsx'):
    """
    Save DataFrame to Excel file with formatting
    
    Args:
        df: pandas DataFrame
        filename: Output filename
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Save to Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # Load workbook for formatting
        wb = load_workbook(filename)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.column == 2:  # Total number column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Adjust column widths
        column_widths = {
            'A': 20,  # UNIT
            'B': 15,  # Total number
            'C': 55,  # DESCRIPTION
            'D': 30,  # Drawing/Ref No
            'E': 18,  # Material
            'F': 20   # Recommended by manufacturer
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 40
        
        # Save formatted workbook
        wb.save(filename)
        print(f"✅ Excel file formatted and saved successfully!")
        
    except ImportError:
        print("⚠️ openpyxl not available for formatting, but file is saved")

def print_statistics(df):
    """Print summary statistics of the data"""
    print("\n" + "="*70)
    print("📊 SUMMARY STATISTICS")
    print("="*70)
    print(f"Total number of part types: {len(df)}")
    print(f"Total quantity of all parts: {df['Total number of identical parts installed'].sum()}")
    print(f"\n🔝 Top 5 parts by quantity:")
    print("-"*70)
    top_5 = df.nlargest(5, 'Total number of identical parts installed')[
        ['Total number of identical parts installed', 'DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)']
    ]
    for idx, row in top_5.iterrows():
        print(f"  • {row['Total number of identical parts installed']:3d} units - {row['DESCRIPTION OF PARTS (Where necessary Give Detail in 2 or 3 Lines Kit & Set Content to be Itemized)'][:50]}")
    print("="*70)

def main():
    """
    Main function to extract and save parts data
    """
    
    print("\n" + "="*70)
    print("🔧 SPARE PARTS LIST EXTRACTOR")
    print("="*70)
    
    # Image path (update this with your actual image file name)
    image_path = 'parts_list_image.png'
    
    print(f"\n📁 Looking for image file: {image_path}")
    
    # Try to extract data from image using OCR
    extracted_data = []
    if os.path.exists(image_path):
        print("🔍 Extracting data from image using OCR...")
        extracted_data = extract_table_data_ocr(image_path)
    
    # If OCR extraction failed or no image, use predefined data
    if not extracted_data:
        if os.path.exists(image_path):
            print("⚠️  OCR extraction did not find data. Using predefined data...")
        else:
            print("ℹ️  Image file not found. Using predefined data from document...")
        extracted_data = get_predefined_data()
    
    # Create DataFrame
    df = pd.DataFrame(extracted_data)
    
    # Save to Excel
    excel_file = 'parts_list_extracted.xlsx'
    print(f"\n💾 Saving to Excel file: {excel_file}")
    save_to_excel(df, excel_file)
    print(f"✅ Excel file saved successfully: {excel_file}")
    
    # Save to CSV
    csv_file = 'parts_list_extracted.csv'
    print(f"\n💾 Saving to CSV file: {csv_file}")
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"✅ CSV file saved successfully: {csv_file}")
    
    # Print statistics
    print_statistics(df)
    
    # Display first 3 records
    print(f"\n📋 Preview of first 3 records:")
    print("-"*70)
    print(df.head(3).to_string(index=False))
    
    print("\n✅ Process completed successfully!")
    print("="*70 + "\n")

if __name__ == "__main__":
    """
    HOW TO RUN THIS SCRIPT:
    
    1. Install required libraries:
       pip install pandas openpyxl pytesseract opencv-python Pillow
    
    2. Install Tesseract OCR (for automatic extraction from images):
       - Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki
       - Linux: sudo apt-get install tesseract-ocr
       - Mac: brew install tesseract
    
    3. Place your image file (parts_list_image.png) in the same directory
    
    4. Run the script:
       python parts_extractor.py
    
    5. Output files will be generated:
       - parts_list_extracted.xlsx (Excel file with formatting)
       - parts_list_extracted.csv (CSV file)
    
    Note: If you don't have an image or Tesseract, the script will still work
    using the predefined data from the document.
    """
    
    main()