
import os
from pathlib import Path
import re
from datetime import datetime
import PyPDF2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
from collections import defaultdict

# تنظیم مسیر Tesseract (در صورت نیاز)
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def parse_date_to_excel(date_str):
    """
    تبدیل تاریخ به فرمت اکسل (dd/mm/yyyy)
    ورودی: 8-Dec-2023 یا 08-Dec-2023
    خروجی: datetime object
    """
    if not date_str:
        return None
    
    try:
        months = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'september': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12
        }
        
        parts = date_str.split('-')
        if len(parts) == 3:
            day = int(parts[0])
            month_name = parts[1].lower()
            year = int(parts[2])
            
            month = months.get(month_name)
            if month:
                return datetime(year, month, day)
    except:
        pass
    
    return None

def extract_text_from_pdf_with_ocr(pdf_path):
    """
    استخراج متن از PDF با استفاده از OCR (برای PDF های اسکن شده)
    """
    try:
        print(f"   🔍 تلاش برای OCR...")
        # تبدیل PDF به تصویر
        images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=300)
        
        if images:
            # OCR روی اولین صفحه
            text = pytesseract.image_to_string(images[0], lang='eng')
            return text
    except Exception as e:
        print(f"   ⚠️ خطا در OCR: {str(e)}")
    
    return ""

def extract_info_from_pdf(pdf_path):
    """
    استخراج اطلاعات از PDF (عادی یا اسکن شده):
    - Doc No
    - Date
    - Report Title (HEAVY CRUDE / LIGHT CRUDE)
    """
    text = ""
    
    try:
        # ابتدا سعی می‌کنیم متن را مستقیم بخوانیم
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            if len(pdf_reader.pages) > 0:
                first_page = pdf_reader.pages[0]
                text = first_page.extract_text()
        
        # اگر متن خالی بود، از OCR استفاده می‌کنیم
        if not text or len(text.strip()) < 50:
            print(f"   ⚠️ PDF اسکن شده است، استفاده از OCR...")
            text = extract_text_from_pdf_with_ocr(pdf_path)
        
        if text:
            print(f"   📄 متن استخراج شده ({len(text)} کاراکتر)")
            print(f"   🔍 نمونه متن: {text[:300].replace(chr(10), ' ')}")
            
            # استخراج Doc No
            # الگوها: Doc No. SJSC-GGNRSP-MOWP-REDA-0001-G00
            patterns_docno = [
                r'Doc\s*No\.?\s*(SJSC-[A-Z]+-[A-Z]+-[A-Z]+-(\d{4})-(G\d{2}))',
                r'(SJSC-[A-Z]+-[A-Z]+-[A-Z]+-(\d{4})-(G\d{2}))',
            ]
            
            doc_no = None
            doc_number = None
            rev = None
            
            for pattern in patterns_docno:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    if len(match.groups()) >= 3:
                        doc_no = match.group(1)
                        doc_number = match.group(2)
                        rev = match.group(3)
                    break
            
            # استخراج تاریخ
            patterns_date = [
                r'Date:\s*([0-9]{1,2}-[A-Za-z]{3,9}-[0-9]{4})',
                r'Date\s*([0-9]{1,2}-[A-Za-z]{3,9}-[0-9]{4})',
                r'([0-9]{1,2}-[A-Za-z]{3,9}-[0-9]{4})',
            ]
            
            date_obj = None
            date_str = None
            
            for pattern in patterns_date:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    date_str = match.group(1)
                    date_obj = parse_date_to_excel(date_str)
                    if date_obj:
                        break
            
            # استخراج عنوان گزارش
            report_title = None
            if 'HEAVY CRUDE' in text.upper():
                report_title = 'DAILY DELIVERY AND ACCEPTANCE REPORT - HEAVY CRUDE'
            elif 'LIGHT CRUDE' in text.upper():
                report_title = 'DAILY DELIVERY AND ACCEPTANCE REPORT - LIGHT CRUDE'
            
            return {
                'doc_no': doc_no,
                'doc_number': doc_number,
                'rev': rev,
                'date': date_obj,
                'date_str': date_str,
                'report_title': report_title
            }
                    
    except Exception as e:
        print(f"   ⚠️ خطا در خواندن PDF: {str(e)}")
    
    return None

def extract_info_from_excel(excel_path):
    """
    استخراج اطلاعات از فایل اکسل:
    - Doc No
    - Date
    - Report Title
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # جستجو در 20 ردیف اول
        doc_no = None
        doc_number = None
        rev = None
        date_obj = None
        date_str = None
        report_title = None
        
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value)
                    
                    # جستجوی Doc No
                    if not doc_no:
                        patterns_docno = [
                            r'Doc\s*No\.?\s*(SJSC-[A-Z]+-[A-Z]+-[A-Z]+-(\d{4})-(G\d{2}))',
                            r'(SJSC-[A-Z]+-[A-Z]+-[A-Z]+-(\d{4})-(G\d{2}))',
                        ]
                        for pattern in patterns_docno:
                            match = re.search(pattern, cell_text, re.IGNORECASE)
                            if match and len(match.groups()) >= 3:
                                doc_no = match.group(1)
                                doc_number = match.group(2)
                                rev = match.group(3)
                                break
                    
                    # جستجوی تاریخ
                    if not date_obj:
                        # بررسی اگر خود سلول datetime است
                        if isinstance(cell.value, datetime):
                            date_obj = cell.value
                            date_str = date_obj.strftime('%d-%b-%Y')
                        else:
                            patterns_date = [
                                r'Date:\s*([0-9]{1,2}-[A-Za-z]{3,9}-[0-9]{4})',
                                r'([0-9]{1,2}-[A-Za-z]{3,9}-[0-9]{4})',
                            ]
                            for pattern in patterns_date:
                                match = re.search(pattern, cell_text, re.IGNORECASE)
                                if match:
                                    date_str = match.group(1)
                                    date_obj = parse_date_to_excel(date_str)
                                    if date_obj:
                                        break
                    
                    # جستجوی عنوان
                    if not report_title:
                        if 'HEAVY CRUDE' in cell_text.upper():
                            report_title = 'DAILY DELIVERY AND ACCEPTANCE REPORT - HEAVY CRUDE'
                        elif 'LIGHT CRUDE' in cell_text.upper():
                            report_title = 'DAILY DELIVERY AND ACCEPTANCE REPORT - LIGHT CRUDE'
        
        return {
            'doc_no': doc_no,
            'doc_number': doc_number,
            'rev': rev,
            'date': date_obj,
            'date_str': date_str,
            'report_title': report_title
        }
    
    except Exception as e:
        print(f"   ⚠️ خطا در خواندن Excel: {str(e)}")
    
    return None

def create_excel_report(files_data, output_path):
    """
    ایجاد گزارش اکسل
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Acceptance Reports"
    
    # استایل‌ها
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # هدرها
    headers = ['ردیف', 'نام فایل اصلی', 'نام فایل جدید', 'عنوان گزارش', 'Doc No', 'شماره', 'REV', 'تاریخ', 'وضعیت']
    ws.append(headers)
    
    # استایل هدر
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # داده‌ها
    for idx, data in enumerate(files_data, start=1):
        date_value = data['date'] if data['date'] else 'N/A'
        
        row = [
            idx,
            data['old_name'],
            data['new_name'],
            data['report_title'] or 'N/A',
            data['doc_no'] or 'N/A',
            data['doc_number'] or 'N/A',
            data['rev'] or 'N/A',
            date_value,
            data['status']
        ]
        ws.append(row)
        
        # استایل ردیف
        row_num = idx + 1
        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # فرمت تاریخ
            if col_idx == 8 and isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY'
    
    # تنظیم عرض ستون‌ها
    column_widths = [8, 40, 45, 50, 35, 12, 8, 15, 20]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # ذخیره فایل
    wb.save(output_path)
    print(f"\n📊 فایل اکسل ایجاد شد: {output_path}")

def rename_files(folder_path):
    """
    تغییر نام فایل‌های PDF و Excel
    """
    print("="*80)
    print("🔄 تغییر نام فایل‌های Acceptance Reports")
    print("="*80)
    print(f"📂 مسیر پوشه: {folder_path}\n")
    
    # پیدا کردن فایل‌ها
    pdf_files = list(Path(folder_path).glob('*.pdf'))
    excel_files = list(Path(folder_path).glob('*.xlsx')) + list(Path(folder_path).glob('*.xls'))
    
    # حذف فایل‌هایی که قبلاً تغییر نام داده شده‌اند
    pdf_files = [f for f in pdf_files if not f.name.startswith('SJSC-GGNRSP-MOWP-REDA-')]
    excel_files = [f for f in excel_files if not f.name.startswith('SJSC-GGNRSP-MOWP-REDA-')]
    
    all_files = pdf_files + excel_files
    
    if not all_files:
        print("❌ هیچ فایلی پیدا نشد!")
        return
    
    print(f"📁 {len(pdf_files)} فایل PDF و {len(excel_files)} فایل Excel پیدا شد\n")
    print("🔍 در حال استخراج اطلاعات...")
    print("-"*80)
    
    # استخراج اطلاعات
    files_data = []
    
    for file_path in all_files:
        print(f"\n📄 پردازش: {file_path.name}")
        
        if file_path.suffix.lower() == '.pdf':
            info = extract_info_from_pdf(file_path)
        else:
            info = extract_info_from_excel(file_path)
        
        if info and info['doc_number'] and info['rev']:
            files_data.append({
                'path': file_path,
                'old_name': file_path.name,
                'info': info,
                'doc_no': info['doc_no'],
                'doc_number': info['doc_number'],
                'rev': info['rev'],
                'date': info['date'],
                'date_str': info['date_str'],
                'report_title': info['report_title'],
                'new_name': None,  # خواهد شد پر
                'status': 'در انتظار'
            })
            
            print(f"   ✅ Doc No: {info['doc_no']}")
            if info['date']:
                print(f"   ✅ Date: {info['date'].strftime('%d/%m/%Y')}")
            print(f"   ✅ Report: {info['report_title']}")
        else:
            files_data.append({
                'path': file_path,
                'old_name': file_path.name,
                'info': None,
                'doc_no': None,
                'doc_number': None,
                'rev': None,
                'date': None,
                'date_str': None,
                'report_title': None,
                'new_name': 'N/A',
                'status': 'خطا - اطلاعات یافت نشد'
            })
            print(f"   ❌ نتوانستیم اطلاعات را پیدا کنیم!")
    
    print("-"*80)
    
    # شناسایی تکراری‌ها و شماره‌گذاری
    print("\n🔢 بررسی تکراری‌ها و تخصیص شماره...")
    
    # گروه‌بندی بر اساس شماره + REV
    groups = defaultdict(list)
    for data in files_data:
        if data['doc_number'] and data['rev']:
            key = f"{data['doc_number']}-{data['rev']}"
            groups[key].append(data)
    
    # تخصیص نام جدید
    for key, group in groups.items():
        if len(group) == 1:
            # فایل یکتا
            data = group[0]
            ext = data['path'].suffix
            data['new_name'] = f"SJSC-GGNRSP-MOWP-REDA-{data['doc_number']}-{data['rev']}{ext}"
        else:
            # فایل‌های تکراری - شماره‌گذاری
            for idx, data in enumerate(group, start=1):
                ext = data['path'].suffix
                data['new_name'] = f"SJSC-GGNRSP-MOWP-REDA-{data['doc_number']}-{data['rev']}_copy{idx}{ext}"
                print(f"   ⚠️ تکراری: {data['doc_number']}-{data['rev']} -> _copy{idx}")
    
    # تغییر نام فایل‌ها
    print("\n🔄 شروع تغییر نام...")
    print("-"*80)
    
    renamed_count = 0
    failed_count = 0
    
    for data in files_data:
        if data['new_name'] and data['new_name'] != 'N/A':
            old_path = data['path']
            new_name = data['new_name']
            new_path = old_path.parent / new_name
            
            if new_path.exists() and new_path != old_path:
                print(f"⚠️ فایل با این نام وجود دارد: {new_name}")
                data['status'] = 'رد شده - نام تکراری'
                failed_count += 1
                continue
            
            try:
                old_path.rename(new_path)
                renamed_count += 1
                data['status'] = '✅ موفق'
                print(f"✅ {old_path.name}")
                print(f"   ➜ {new_name}")
            except Exception as e:
                print(f"❌ خطا در تغییر نام: {str(e)}")
                data['status'] = f'❌ خطا: {str(e)}'
                failed_count += 1
    
    print("-"*80)
    
    # ایجاد گزارش اکسل
    excel_path = Path(folder_path) / f"Acceptance_Rename_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    create_excel_report(files_data, excel_path)
    
    # خلاصه
    print(f"\n📊 نتیجه:")
    print(f"   ✅ تعداد فایل‌های تغییر نام داده شده: {renamed_count}")
    print(f"   ❌ تعداد فایل‌های ناموفق: {failed_count}")
    print("="*80)

def main():
    """
    تابع اصلی
    """
    FOLDER_PATH = r"D:\Sepher_Pasargad\works\Production\Acceptance"
    
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ خطا: پوشه پیدا نشد!")
        print(f"مسیر: {FOLDER_PATH}")
        return
    
    print("\n⚠️ هشدار: این عملیات نام فایل‌های PDF و Excel را تغییر می‌دهد!")
    print("⚠️ برای PDF های اسکن شده از OCR استفاده می‌شود (نیاز به Tesseract)")
    print("\nآیا مطمئن هستید؟ (y/n): ", end='')
    
    confirmation = input().lower()
    if confirmation != 'y':
        print("❌ عملیات لغو شد.")
        return
    
    rename_files(FOLDER_PATH)
    
    print("\n✨ کار تمام شد!")

if __name__ == "__main__":
    main()