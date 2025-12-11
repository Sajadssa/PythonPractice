import os
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import win32com.client
import pythoncom
from PyPDF2 import PdfMerger
import shutil

class ExcelToPdfProcessor:
    def __init__(self, directory_path):
        self.directory_path = directory_path
        self.results = []
        self.temp_pdf_folder = os.path.join(directory_path, "temp_pdfs")
        
        # ایجاد پوشه موقت
        if not os.path.exists(self.temp_pdf_folder):
            os.makedirs(self.temp_pdf_folder)
    
    def extract_info_from_excel(self, excel_path):
        """
        استخراج Doc No و Date از فایل اکسل
        """
        doc_no = None
        date = None
        number = None
        rev = None
        
        try:
            # خواندن با openpyxl برای دسترسی بهتر به سلول‌ها
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # جستجو در تمام سلول‌های چند ردیف اول
            text = ""
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    if cell:
                        text += str(cell) + " "
            
            wb.close()
            
            print(f"  🔍 متن استخراج شده: {text[:200]}...")
            
            # جستجوی Doc No با فرمت دقیق: SJSC-GGNRSP-PDPE-REDH-XXXX-GXX
            doc_pattern = r'(SJSC-GGNRSP-[A-Z]+-[A-Z]+-(\d{4})-(G\d{2}))'
            match = re.search(doc_pattern, text, re.IGNORECASE)
            
            if match:
                doc_no = match.group(1)
                number = match.group(2)  # 4 رقم وسط
                rev = match.group(3)      # Gxx
                print(f"  ✅ Doc No پیدا شد: {doc_no}")
                print(f"  ✅ Number: {number}, Rev: {rev}")
            else:
                # تلاش دیگر برای Doc No
                alt_pattern = r'Doc\s*No\.?\s*:?\s*([A-Z0-9\-]+)'
                match2 = re.search(alt_pattern, text, re.IGNORECASE)
                if match2:
                    doc_no = match2.group(1)
                    # استخراج number و rev از doc_no
                    parts = doc_no.split('-')
                    for i, part in enumerate(parts):
                        if re.match(r'\d{4}', part):
                            number = part
                            if i + 1 < len(parts):
                                rev_part = parts[i + 1]
                                if re.match(r'G?\d{2}', rev_part):
                                    rev = 'G' + re.sub(r'[^0-9]', '', rev_part).zfill(2)
                            break
            
            # جستجوی Date با فرمت: 1-Aug-2024
            date_patterns = [
                r'Date\s*:?\s*(\d{1,2}-[A-Za-z]{3}-\d{4})',
                r'Date\s*:?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})',
                r'(\d{1,2}-[A-Za-z]{3}-\d{4})',
                r'(\d{1,2}/\d{1,2}/\d{4})',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    date = match.group(1).strip()
                    print(f"  ✅ Date پیدا شد: {date}")
                    break
            
        except Exception as e:
            print(f"  ⚠️  خطا در خواندن اکسل: {e}")
        
        return {
            'doc_no': doc_no,
            'number': number,
            'rev': rev,
            'date': date
        }
    
    def excel_to_pdf(self, excel_path, pdf_path):
        """
        تبدیل فایل Excel به PDF
        """
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # باز کردن فایل
            wb = excel.Workbooks.Open(os.path.abspath(excel_path))
            
            # تنظیمات چاپ برای کیفیت بهتر
            ws = wb.Worksheets(1)
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False
            
            # تبدیل به PDF
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            
            # بستن
            wb.Close(False)
            excel.Quit()
            pythoncom.CoUninitialize()
            
            return True
            
        except Exception as e:
            print(f"  ❌ خطا در تبدیل به PDF: {e}")
            try:
                excel.Quit()
                pythoncom.CoUninitialize()
            except:
                pass
            return False
    
    def identify_file_type(self, filename):
        """
        شناسایی نوع فایل (Heavy یا Light)
        """
        filename_lower = filename.lower()
        if 'heavy' in filename_lower:
            return 'Heavy Crude'
        elif 'light' in filename_lower:
            return 'Light Crude'
        else:
            return 'Unknown'
    
    def process_excel_files(self):
        """
        پردازش تمام فایل‌های Excel
        """
        # دریافت لیست فایل‌های Excel
        excel_files = [f for f in os.listdir(self.directory_path) 
                       if f.lower().endswith(('.xlsx', '.xls')) 
                       and not f.startswith('~$')  # فایل‌های موقت اکسل را نادیده بگیر
                       and not f.startswith('Excel_to_PDF_Report')]  # گزارش قبلی را نادیده بگیر
        
        if not excel_files:
            print("❌ هیچ فایل Excel یافت نشد!")
            return
        
        print(f"📁 تعداد {len(excel_files)} فایل Excel یافت شد.\n")
        print("="*80)
        
        # مرحله 1: تبدیل به PDF و استخراج اطلاعات
        print("\n🔄 مرحله 1: تبدیل Excel به PDF و استخراج اطلاعات...\n")
        
        file_info_dict = {}  # {number: [list of files with same number]}
        
        for idx, excel_file in enumerate(excel_files, 1):
            print(f"[{idx}/{len(excel_files)}] 📄 {excel_file}")
            
            excel_path = os.path.join(self.directory_path, excel_file)
            
            # شناسایی نوع فایل
            file_type = self.identify_file_type(excel_file)
            print(f"  📋 نوع: {file_type}")
            
            # استخراج اطلاعات
            info = self.extract_info_from_excel(excel_path)
            
            if not info['number'] or not info['rev']:
                print(f"  ⚠️  اطلاعات ناقص - Number: {info['number']}, Rev: {info['rev']}\n")
                self.results.append({
                    'نام فایل Excel': excel_file,
                    'نوع گزارش': file_type,
                    'Doc No اصلی': info['doc_no'] or 'نامشخص',
                    'Number': info['number'] or 'نامشخص',
                    'Rev': info['rev'] or 'نامشخص',
                    'تاریخ': info['date'] or 'نامشخص',
                    'نام فایل PDF نهایی': 'ناموفق',
                    'وضعیت': 'ناموفق - اطلاعات ناقص'
                })
                continue
            
            # تبدیل به PDF
            temp_pdf_name = f"temp_{info['number']}_{file_type.replace(' ', '_')}_{idx}.pdf"
            temp_pdf_path = os.path.join(self.temp_pdf_folder, temp_pdf_name)
            
            print(f"  🔄 در حال تبدیل به PDF...")
            if self.excel_to_pdf(excel_path, temp_pdf_path):
                print(f"  ✅ تبدیل موفق\n")
                
                # ذخیره اطلاعات
                if info['number'] not in file_info_dict:
                    file_info_dict[info['number']] = []
                
                file_info_dict[info['number']].append({
                    'original_name': excel_file,
                    'temp_pdf': temp_pdf_path,
                    'type': file_type,
                    'info': info
                })
            else:
                print(f"  ❌ تبدیل ناموفق\n")
                self.results.append({
                    'نام فایل Excel': excel_file,
                    'نوع گزارش': file_type,
                    'Doc No اصلی': info['doc_no'],
                    'Number': info['number'],
                    'Rev': info['rev'],
                    'تاریخ': info['date'],
                    'نام فایل PDF نهایی': 'ناموفق',
                    'وضعیت': 'ناموفق - خطا در تبدیل PDF'
                })
        
        # مرحله 2: ادغام و تغییر نام
        print("\n" + "="*80)
        print("🔄 مرحله 2: ادغام فایل‌های با Number یکسان و تغییر نام...\n")
        
        for number, files in file_info_dict.items():
            print(f"📊 Number: {number} - تعداد فایل: {len(files)}")
            
            # نمایش فایل‌ها
            for f in files:
                print(f"  • {f['original_name']} ({f['type']})")
            
            # نام فایل نهایی با فرمت صحیح
            rev = files[0]['info']['rev']
            final_pdf_name = f"SJSC-GGNRSP-MOWP-REDA-{number}-{rev}.pdf"
            final_pdf_path = os.path.join(self.directory_path, final_pdf_name)
            
            try:
                if len(files) > 1:
                    # ادغام چند فایل
                    print(f"  🔗 ادغام {len(files)} فایل...")
                    merger = PdfMerger()
                    
                    # مرتب‌سازی: Heavy اول، بعد Light
                    files_sorted = sorted(files, key=lambda x: 0 if 'Heavy' in x['type'] else 1)
                    
                    for f in files_sorted:
                        merger.append(f['temp_pdf'])
                    
                    merger.write(final_pdf_path)
                    merger.close()
                    
                    print(f"  ✅ ادغام موفق و ذخیره شد: {final_pdf_name}\n")
                    
                    # ثبت نتیجه برای هر فایل
                    for f in files:
                        self.results.append({
                            'نام فایل Excel': f['original_name'],
                            'نوع گزارش': f['type'],
                            'Doc No اصلی': f['info']['doc_no'],
                            'Number': f['info']['number'],
                            'Rev': f['info']['rev'],
                            'تاریخ': f['info']['date'],
                            'نام فایل PDF نهایی': final_pdf_name,
                            'وضعیت': f'موفق - ادغام شده با {len(files)} فایل'
                        })
                else:
                    # فقط یک فایل - کپی مستقیم
                    print(f"  📋 تنها یک فایل وجود دارد - کپی مستقیم...")
                    shutil.copy2(files[0]['temp_pdf'], final_pdf_path)
                    print(f"  ✅ ذخیره شد: {final_pdf_name}\n")
                    
                    self.results.append({
                        'نام فایل Excel': files[0]['original_name'],
                        'نوع گزارش': files[0]['type'],
                        'Doc No اصلی': files[0]['info']['doc_no'],
                        'Number': files[0]['info']['number'],
                        'Rev': files[0]['info']['rev'],
                        'تاریخ': files[0]['info']['date'],
                        'نام فایل PDF نهایی': final_pdf_name,
                        'وضعیت': 'موفق - فایل واحد'
                    })
                    
            except Exception as e:
                print(f"  ❌ خطا در ادغام/کپی: {e}\n")
                for f in files:
                    self.results.append({
                        'نام فایل Excel': f['original_name'],
                        'نوع گزارش': f['type'],
                        'Doc No اصلی': f['info']['doc_no'],
                        'Number': f['info']['number'],
                        'Rev': f['info']['rev'],
                        'تاریخ': f['info']['date'],
                        'نام فایل PDF نهایی': 'ناموفق',
                        'وضعیت': f'ناموفق - خطا: {str(e)[:50]}'
                    })
        
        # پاک کردن پوشه موقت
        print("\n🗑️  حذف فایل‌های موقت...")
        try:
            shutil.rmtree(self.temp_pdf_folder)
            print("✅ فایل‌های موقت حذف شدند")
        except Exception as e:
            print(f"⚠️  خطا در حذف فایل‌های موقت: {e}")
    
    def save_report(self):
        """
        ذخیره گزارش در Excel
        """
        if not self.results:
            print("\n⚠️  هیچ نتیجه‌ای برای ذخیره وجود ندارد!")
            return
        
        df = pd.DataFrame(self.results)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"Excel_to_PDF_Report_{timestamp}.xlsx"
        excel_path = os.path.join(self.directory_path, excel_filename)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='گزارش تبدیل و ادغام', index=False)
            
            # تنظیم عرض ستون‌ها
            worksheet = writer.sheets['گزارش تبدیل و ادغام']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # خلاصه
            successful = len([r for r in self.results if 'موفق' in r['وضعیت']])
            failed = len([r for r in self.results if 'ناموفق' in r['وضعیت']])
            
            # تعداد فایل‌های PDF نهایی ایجاد شده
            unique_pdfs = len(set([r['نام فایل PDF نهایی'] for r in self.results if r['نام فایل PDF نهایی'] != 'ناموفق']))
            
            summary_data = {
                'شرح': [
                    'تعداد کل فایل‌های Excel',
                    'تعداد تبدیل موفق',
                    'تعداد ناموفق',
                    'تعداد فایل‌های PDF نهایی',
                    'درصد موفقیت',
                    'تاریخ و زمان'
                ],
                'مقدار': [
                    len(self.results),
                    successful,
                    failed,
                    unique_pdfs,
                    f"{(successful/len(self.results)*100):.1f}%" if self.results else "0%",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='خلاصه', index=False)
            
            worksheet_summary = writer.sheets['خلاصه']
            worksheet_summary.column_dimensions['A'].width = 35
            worksheet_summary.column_dimensions['B'].width = 30
        
        print(f"\n📊 گزارش Excel ذخیره شد:")
        print(f"   {excel_path}")
        print(f"\n📈 خلاصه نتایج:")
        print(f"   ✅ موفق: {successful}")
        print(f"   ❌ ناموفق: {failed}")
        print(f"   📄 فایل‌های PDF نهایی: {unique_pdfs}")
        
        return excel_path


def main():
    folder_path = r"D:\Sepher_Pasargad\works\Production\Daily_Acceptance"
    
    print("🚀 شروع پردازش فایل‌های Excel...")
    print("="*80)
    print("📋 مراحل:")
    print("   1️⃣  تبدیل Excel به PDF")
    print("   2️⃣  استخراج Doc No و Date")
    print("   3️⃣  ادغام فایل‌های Heavy و Light با Number یکسان")
    print("   4️⃣  تغییر نام به فرمت: SJSC-GGNRSP-MOWP-REDA-XXXX-GXX")
    print("   5️⃣  ایجاد گزارش Excel")
    print("="*80)
    
    processor = ExcelToPdfProcessor(folder_path)
    
    # پردازش فایل‌ها
    processor.process_excel_files()
    
    # ذخیره گزارش
    print("\n" + "="*80)
    processor.save_report()
    
    print("\n✨ پردازش کامل شد!")
    print("="*80)


if __name__ == "__main__":
    main()