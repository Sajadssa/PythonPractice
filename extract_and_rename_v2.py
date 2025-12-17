#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

try:
    import pdfplumber
    PDF_OK = True
except:
    PDF_OK = False

try:
    from docx import Document
    DOCX_OK = True
except:
    DOCX_OK = False


def extract_pdf(path):
    if not PDF_OK:
        return None, None
    try:
        with pdfplumber.open(path) as pdf:
            text = pdf.pages[0].extract_text()
            
            # Extract Sequence Number from header table
            seq = re.search(r'Sequence\s+Number\s+(\d+)', text, re.I)
            seq_num = seq.group(1).zfill(4) if seq else None
            
            # Extract Revision from header table
            rev = re.search(r'Revision\s+([GG]\d+)', text, re.I)
            revision = rev.group(1).upper() if rev else None
            
            # Extract Date from bottom table
            date = re.search(r'(\d{1,2}[-/][A-Za-z]{3}[-/]\d{4})', text)
            date_str = date.group(1) if date else None
            
            return seq_num, revision, date_str
    except:
        return None, None, None


def extract_docx(path):
    if not DOCX_OK:
        return None, None, None
    try:
        doc = Document(path)
        text = ""
        for table in doc.tables[:2]:
            for row in table.rows:
                for cell in row.cells:
                    text += cell.text + " | "
        
        # Extract Sequence Number
        seq = re.search(r'Sequence\s+Number\s+(\d+)', text, re.I)
        seq_num = seq.group(1).zfill(4) if seq else None
        
        # Extract Revision
        rev = re.search(r'Revision\s+([GG]\d+)', text, re.I)
        revision = rev.group(1).upper() if rev else None
        
        # Extract Date
        date = re.search(r'(\d{1,2}[-/][A-Za-z]{3}[-/]\d{4})', text)
        date_str = date.group(1) if date else None
        
        return seq_num, revision, date_str
    except:
        return None, None, None


def process_files(directory):
    results = []
    
    for file in Path(directory).iterdir():
        if file.name.startswith('~$'):
            continue
            
        ext = file.suffix.lower()
        if ext not in ['.pdf', '.docx', '.doc']:
            continue
        
        print(f"Processing: {file.name}")
        
        if ext == '.pdf':
            seq, rev, date = extract_pdf(file)
        else:
            seq, rev, date = extract_docx(file)
        
        if seq and rev:
            new_name = f"SJSC-GGNRSP-MADR-REMO-{seq}-{rev}{ext}"
            results.append({
                'old_name': file.name,
                'new_name': new_name,
                'date': date or 'N/A',
                'old_path': file
            })
            print(f"  -> {new_name}")
        else:
            print(f"  FAILED: Could not extract data")
    
    return results


def create_excel(results, output):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Renamed Files"
    
    headers = ['Original Name', 'New Name', 'Date']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    for item in results:
        ws.append([item['old_name'], item['new_name'], item['date']])
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    
    wb.save(output)
    print(f"\nExcel created: {output}")


def rename_files(results):
    for item in results:
        old = item['old_path']
        new = old.parent / item['new_name']
        
        if new.exists():
            print(f"EXISTS: {new.name}")
            continue
        
        try:
            old.rename(new)
            print(f"RENAMED: {old.name} -> {new.name}")
        except Exception as e:
            print(f"ERROR: {e}")


def main():
    DIR = r"D:\Sepher_Pasargad\works\Maintenace\Maintenance Report\All_Extracted\monthly"
    EXCEL = r"D:\Sepher_Pasargad\works\Maintenace\Maintenance Report\All_Extracted\renamed_files.xlsx"
    
    print("Processing files...\n")
    results = process_files(DIR)
    
    if not results:
        print("\nNo files processed!")
        return
    
    print(f"\n{len(results)} files processed")
    
    create_excel(results, EXCEL)
    
    ans = input("\nRename files? (yes/no): ").lower()
    if ans in ['yes', 'y']:
        rename_files(results)


if __name__ == "__main__":
    main()
