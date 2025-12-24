#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
اسکریپت تغییر نام فایل‌های Daily Production Operation Report
استخراج Number و Revision از Doc. No.
"""

import os
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

try:
    from PyPDF2 import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False


class DailyReportRenamer:
    def __init__(self, source_directory):
        """
        مقداردهی اولیه
        
        Args:
            source_directory: مسیر پوشه حاوی فایل‌های Excel و PDF
        """
        self.source_directory = Path(source_directory)
        self.results = []
        
    def extract_from_excel(self, file_path):
        """
        استخراج Number و Revision از Doc. No. در Excel
        با پشتیبانی کامل از merged cells
        
        Args:
            file_path: مسیر فایل Excel
            
        Returns:
            dict: دیکشنری حاوی number و revision
        """
        data = {
            'number': None,
            'revision': None,
            'date': None
        }
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            
            # جستجو در تمام سلول‌ها (حتی merged)
            for row in range(1, 50):  # 50 ردیف اول
                for col in range(1, 20):  # 20 ستون اول
                    try:
                        cell = ws.cell(row, col)
                        cell_value = cell.value
                        
                        if cell_value and isinstance(cell_value, str):
                            # جستجو برای Doc. No. با الگوهای مختلف
                            if 'Doc. No.' in cell_value or 'Doc.No.' in cell_value or 'Document No' in cell_value:
                                # الگوهای مختلف:
                                # 1. SJSC-GGNRSP-PDOP-REDA-0060-G00
                                # 2. SJSC-GGNRSP-EPWC-REDA-0060-G00
                                patterns = [
                                    r'REDA-(\d{4})-(G\d{2})',
                                    r'REDA-(\d{3,4})-(G\d{1,2})',
                                    r'REDA[-_](\d{4})[-_](G\d{2})',
                                ]
                                
                                for pattern in patterns:
                                    match = re.search(pattern, cell_value)
                                    if match:
                                        data['number'] = match.group(1).zfill(4)
                                        data['revision'] = match.group(2)
                                        if len(data['revision']) == 2:  # اگه G0 باشه
                                            data['revision'] = data['revision'] + '0'
                                        break
                                
                                if data['number']:
                                    break
                            
                            # جستجو برای Date
                            if not data['date']:
                                # الگوهای مختلف تاریخ
                                date_patterns = [
                                    r'Date:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})',
                                    r'\b(\d{1,2}-[A-Za-z]{3}-\d{4})\b',
                                    r'(\d{1,2}/\d{1,2}/\d{4})',
                                ]
                                
                                for date_pattern in date_patterns:
                                    date_match = re.search(date_pattern, cell_value)
                                    if date_match:
                                        data['date'] = date_match.group(1)
                                        break
                    
                    except:
                        continue
                
                if data['number']:
                    break
            
            wb.close()
            
        except Exception as e:
            print(f"      خطا در خواندن Excel: {str(e)}")
        
        return data
    
    def extract_from_pdf(self, file_path):
        """
        استخراج Number و Revision از Doc. No. در PDF
        با الگوهای گسترده‌تر
        
        Args:
            file_path: مسیر فایل PDF
            
        Returns:
            dict: دیکشنری حاوی number و revision
        """
        data = {
            'number': None,
            'revision': None,
            'date': None
        }
        
        if not PDF_SUPPORT:
            return data
        
        try:
            reader = PdfReader(file_path)
            
            # جستجو در همه صفحات (نه فقط اولی)
            for page_num in range(min(3, len(reader.pages))):  # 3 صفحه اول
                page = reader.pages[page_num]
                text = page.extract_text()
                
                if not data['number']:
                    # الگوهای مختلف برای Doc. No.
                    patterns = [
                        r'REDA-(\d{4})-(G\d{2})',
                        r'REDA-(\d{3,4})-(G\d{1,2})',
                        r'REDA[-_\s](\d{4})[-_\s](G\d{2})',
                        r'Doc\.\s*No\.?\s*:?\s*SJSC-GGNRSP-[A-Z]+-REDA-(\d{4})-(G\d{2})',
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, text, re.IGNORECASE)
                        if match:
                            data['number'] = match.group(1).zfill(4)
                            data['revision'] = match.group(2)
                            if len(data['revision']) == 2:
                                data['revision'] = data['revision'] + '0'
                            break
                
                if not data['date']:
                    # الگوهای مختلف تاریخ
                    date_patterns = [
                        r'Date:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})',
                        r'\b(\d{1,2}-[A-Za-z]{3}-\d{4})\b',
                        r'(\d{1,2}/\d{1,2}/\d{4})',
                    ]
                    
                    for date_pattern in date_patterns:
                        dates = re.findall(date_pattern, text)
                        if dates:
                            data['date'] = dates[0]
                            break
                
                # اگه هر دو پیدا شدن، break
                if data['number'] and data['date']:
                    break
        
        except Exception as e:
            print(f"      خطا در خواندن PDF: {str(e)}")
        
        return data
    
    def extract_from_filename(self, filename):
        """
        استخراج Number و Revision از نام فایل (fallback)
        
        Args:
            filename: نام فایل
            
        Returns:
            dict: دیکشنری حاوی number و revision
        """
        data = {
            'number': None,
            'revision': None
        }
        
        # الگوهای ممکن در نام فایل
        patterns = [
            r'REDA-(\d{4})-(G\d{2})',
            r'REDA-(\d{3,4})-(G\d{1,2})',
            r'(\d{4})-(G\d{2})',
            r'_(\d{4})_',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                if len(match.groups()) >= 2:
                    data['number'] = match.group(1).zfill(4)
                    data['revision'] = match.group(2)
                elif len(match.groups()) == 1:
                    data['number'] = match.group(1).zfill(4)
                break
        
        return data
    
    def generate_new_filename(self, number, revision, extension):
        """
        تولید نام جدید فایل
        
        Args:
            number: شماره (مثلاً 0060)
            revision: ویرایش (مثلاً G00)
            extension: پسوند (.xlsx یا .pdf)
            
        Returns:
            str: نام جدید فایل
        """
        if not number:
            number = "0000"
        if not revision:
            revision = "G00"
        
        # فرمت: SJSC-GGNRSP-EPWC-REDA-[Number]-[Revision]
        new_name = f"SJSC-GGNRSP-EPWC-REDA-{number}-{revision}{extension}"
        return new_name
    
    def process_files(self, rename_files=False):
        """
        پردازش فایل‌های Excel و PDF
        
        Args:
            rename_files: آیا فایل‌ها تغییر نام پیدا کنند؟
            
        Returns:
            list: لیست نتایج
        """
        # پیدا کردن فایل‌های Excel
        excel_files = list(self.source_directory.glob("*.xlsx"))
        excel_files.extend(list(self.source_directory.glob("*.xls")))
        
        # پیدا کردن فایل‌های PDF
        pdf_files = list(self.source_directory.glob("*.pdf"))
        
        print(f"تعداد {len(excel_files)} فایل Excel پیدا شد.")
        print(f"تعداد {len(pdf_files)} فایل PDF پیدا شد.\n")
        
        # پردازش Excel
        for file_path in excel_files:
            if file_path.name.startswith('~$'):
                continue
            
            print(f"📊 در حال پردازش Excel: {file_path.name}")
            
            try:
                data = self.extract_from_excel(file_path)
                
                # اگه پیدا نشد، از نام فایل استخراج کن
                if not data['number']:
                    filename_data = self.extract_from_filename(file_path.name)
                    if filename_data['number']:
                        data['number'] = filename_data['number']
                        data['revision'] = filename_data['revision'] or data['revision']
                        print(f"  ℹ️  استخراج از نام فایل")
                
                print(f"  Number: {data['number']}")
                print(f"  Revision: {data['revision']}")
                print(f"  Date: {data['date']}")
                
                new_filename = self.generate_new_filename(
                    data['number'], 
                    data['revision'],
                    '.xlsx'
                )
                
                print(f"  نام جدید: {new_filename}\n")
                
                result = {
                    'file_type': 'Excel',
                    'original_name': file_path.name,
                    'new_name': new_filename,
                    'number': data['number'] or 'N/A',
                    'revision': data['revision'] or 'N/A',
                    'date': data['date'] or 'N/A',
                    'status': 'پردازش شد'
                }
                
                if rename_files:
                    new_path = file_path.parent / new_filename
                    if not new_path.exists():
                        file_path.rename(new_path)
                        result['status'] = 'تغییر نام داده شد'
                    else:
                        result['status'] = 'فایل با این نام وجود دارد'
                
                self.results.append(result)
                
            except Exception as e:
                print(f"  خطا: {str(e)}\n")
                self.results.append({
                    'file_type': 'Excel',
                    'original_name': file_path.name,
                    'new_name': 'خطا',
                    'number': 'N/A',
                    'revision': 'N/A',
                    'date': 'N/A',
                    'status': f'خطا: {str(e)}'
                })
        
        # پردازش PDF
        for file_path in pdf_files:
            print(f"📕 در حال پردازش PDF: {file_path.name}")
            
            try:
                data = self.extract_from_pdf(file_path)
                
                # اگه پیدا نشد، از نام فایل استخراج کن
                if not data['number']:
                    filename_data = self.extract_from_filename(file_path.name)
                    if filename_data['number']:
                        data['number'] = filename_data['number']
                        data['revision'] = filename_data['revision'] or data['revision']
                        print(f"  ℹ️  استخراج از نام فایل")
                
                print(f"  Number: {data['number']}")
                print(f"  Revision: {data['revision']}")
                print(f"  Date: {data['date']}")
                
                new_filename = self.generate_new_filename(
                    data['number'], 
                    data['revision'],
                    '.pdf'
                )
                
                print(f"  نام جدید: {new_filename}\n")
                
                result = {
                    'file_type': 'PDF',
                    'original_name': file_path.name,
                    'new_name': new_filename,
                    'number': data['number'] or 'N/A',
                    'revision': data['revision'] or 'N/A',
                    'date': data['date'] or 'N/A',
                    'status': 'پردازش شد'
                }
                
                if rename_files:
                    new_path = file_path.parent / new_filename
                    if not new_path.exists():
                        file_path.rename(new_path)
                        result['status'] = 'تغییر نام داده شد'
                    else:
                        result['status'] = 'فایل با این نام وجود دارد'
                
                self.results.append(result)
                
            except Exception as e:
                print(f"  خطا: {str(e)}\n")
                self.results.append({
                    'file_type': 'PDF',
                    'original_name': file_path.name,
                    'new_name': 'خطا',
                    'number': 'N/A',
                    'revision': 'N/A',
                    'date': 'N/A',
                    'status': f'خطا: {str(e)}'
                })
        
        return self.results
    
    def create_excel_report(self, output_path):
        """
        تولید گزارش Excel
        
        Args:
            output_path: مسیر فایل خروجی Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش Daily Reports"
        
        # تنظیمات سبک
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # هدر جدول
        headers = ['ردیف', 'نوع', 'نام اصلی فایل', 'نام جدید فایل', 'Number', 'Revision', 'Date', 'وضعیت']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # داده‌ها
        for row_idx, result in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1).border = border
            ws.cell(row=row_idx, column=2, value=result['file_type']).border = border
            ws.cell(row=row_idx, column=3, value=result['original_name']).border = border
            ws.cell(row=row_idx, column=4, value=result['new_name']).border = border
            ws.cell(row=row_idx, column=5, value=result['number']).border = border
            ws.cell(row=row_idx, column=6, value=result['revision']).border = border
            ws.cell(row=row_idx, column=7, value=result['date']).border = border
            ws.cell(row=row_idx, column=8, value=result['status']).border = border
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        # ذخیره فایل
        wb.save(output_path)
        print(f"\n✓ گزارش Excel در مسیر زیر ذخیره شد:")
        print(f"  {output_path}")


def main():
    """
    تابع اصلی برنامه
    """
    print("=" * 70)
    print("برنامه تغییر نام فایل‌های Daily Production Operation Report")
    print("=" * 70)
    print()
    
    # مسیر پوشه فایل‌ها
    source_dir = r"D:\Sepher_Pasargad\works\Production\02-Operation Daily Activity Report"
    
    # بررسی وجود پوشه
    if not os.path.exists(source_dir):
        print(f"خطا: پوشه {source_dir} یافت نشد!")
        print("لطفاً مسیر را در کد بررسی کنید.")
        return
    
    # ایجاد شیء پردازشگر
    renamer = DailyReportRenamer(source_dir)
    
    # پردازش فایل‌ها (بدون تغییر نام واقعی)
    print("در حال پردازش فایل‌ها...")
    print("-" * 70)
    results = renamer.process_files(rename_files=False)
    
    # تولید گزارش Excel
    output_excel = os.path.join(source_dir, f"گزارش_Daily_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    renamer.create_excel_report(output_excel)
    
    # خلاصه نتایج
    print("\n" + "=" * 70)
    print("خلاصه نتایج:")
    print(f"  تعداد کل فایل‌ها: {len(results)}")
    print(f"  Excel: {sum(1 for r in results if r['file_type'] == 'Excel')}")
    print(f"  PDF: {sum(1 for r in results if r['file_type'] == 'PDF')}")
    print(f"  موفق: {sum(1 for r in results if 'خطا' not in r['status'])}")
    print(f"  خطا: {sum(1 for r in results if 'خطا' in r['status'])}")
    print("=" * 70)
    
    # سوال برای تغییر نام واقعی
    print("\nآیا می‌خواهید فایل‌ها واقعاً تغییر نام پیدا کنند?")
    print("توجه: این عملیات قابل بازگشت نیست!")
    choice = input("برای تغییر نام 'yes' وارد کنید: ")
    
    if choice.lower() == 'yes':
        print("\nدر حال تغییر نام فایل‌ها...")
        renamer.results = []
        results = renamer.process_files(rename_files=True)
        
        # تولید گزارش جدید
        output_excel_final = os.path.join(source_dir, f"گزارش_نهایی_Daily_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        renamer.create_excel_report(output_excel_final)
        print("\n✓ فایل‌ها با موفقیت تغییر نام یافتند!")


if __name__ == "__main__":
    main()
