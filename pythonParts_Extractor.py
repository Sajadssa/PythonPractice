import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# داده‌های استخراج شده از تصویر - با تعداد دقیق
data = {
    'Row': list(range(1, 47)),  # 46 ردیف
    'Qty': [
        38, 57, 26, 19, 31, 38, 26, 14, 19, 71, 
        38, 19, 11, 11, 6, 11, 6, 19, 19, 19, 
        19, 19, 19, 19, 57, 76, 38, 114, 38, 19, 
        38, 19, 171, 144, 133, 171, 114, 114, 38, 38, 
        76, 38, 114, 114, 114, 114
    ],
    'Description of Parts': [
        'Pressure Gauge (0 - 250 Bar)',
        'Pressure Gauge (0 - 400 Bar)',
        'Pressure Gauge (0 - 10,000 PSI)',
        'Pressure Transmitter (0 - 25 Bar)',
        'Pressure Transmitter (0 - 300 Bar)',
        'Pressure Transmitter (0 - 600 Bar)',
        'Pressure Transmitter (0 - 1000 Bar)',
        'High Pressure Filter (1/4", 10,000 Psi)',
        'Repair Kit for High Pressure Filter',
        'Element for High Pressure Filter',
        'Return Filter',
        'Strainer Y Type',
        'Interface Valve',
        'Hydraulic Regulator (1/4", 0 - 150 PSI)',
        'Hydraulic Regulator (1/2", 50 - 6000 PSI)',
        'Repair Kit for Hydraulic Regulator 6000 Psi & 150 Psi',
        'Hydraulic Regulator (1/4", 200 - 10,000 PSI)',
        'Hydraulic Regulator (1/2", 200 - 10,000 PSI)',
        'Repair Kit for Hydraulic Regulator (200 - 10,000 PSI)',
        'Pilot Valve (1/4", 10 Bar)',
        'Flow Line Pilot Valve (1/4", 6000 Psi)',
        'Flow Control Valve (1/4", 1000 Psi)',
        'Push Button (ESD Line Charge)',
        'Push Button (Emergency Stop)',
        'Level Transmitter',
        'Level Switch',
        'Level Gauge (Gauge Glass)',
        'Complete Assembled Electro Pump (SSSV & SSV)',
        'Complete Assembled Electro Pump (HIPPS & ESDV)',
        'Pump (SSSV & SSV)',
        'Pump (HIPPS & ESDV)',
        'Electro Motor',
        'Coupling for Electro Pump',
        'Hand Pump (10,000 Psi)',
        'Hand Pump (5000 Psi)',
        'Power Line Controller',
        'Signal Lamp (Green)',
        'Signal Lamp (Red)',
        'Signal Lamp (Yellow)',
        'Selector Switch (2 Position)',
        'Selector Switch (3 Position)',
        'Contactor',
        'MCB 3 Phase (2 A)',
        'MCB 1 Phase (2 A)',
        'MCCB (5 A)',
        'MPCB (2.8-4 A)'
    ],
    'Material (see Note 4 above)': [
        'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316',
        'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316',
        'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316',
        'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316',
        'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316', 'SS 316',
        'Electrical Component', 'Electrical Component', 'Electrical Component',
        'Electrical Component', 'Electrical Component', 'Electrical Component',
        'Electrical Component', 'Electrical Component', 'Electrical Component',
        'Electrical Component', 'Electrical Component'
    ],
    'Recommended by Manufacturer for 19 MPD': [
        '4 No.', '5 No.', '3 No.', '2 No.', '2 No.', '2 No.', '2 No.',
        '5 No.', '5 No.', '5 No.', '2 No.', '3 No.', '4 No.', '3 No.',
        '3 No.', '5 No.', '3 No.', '3 No.', '2 No.', '4 No.', '4 No.',
        '4 No.', '1 No.', '1 No.', '1 No.', '1 No.', '4 No.', '1 No.',
        '1 No.', '1 No.', '1 No.', '2 No.', '2 No.', '1 No.', '1 No.',
        '3 No.', '5 No.', '10 No.', '8 No.', '2 No.', '8 No.', '5 No.',
        '3 No.', '3 No.', '3 No.', '6 No.'
    ]
}

# بررسی و نمایش تعداد عناصر
print("="*60)
print("بررسی تعداد عناصر در هر ستون:")
print("="*60)
for key, value in data.items():
    print(f"{key}: {len(value)} آیتم")

# بررسی یکسان بودن تعداد
lengths = [len(v) for v in data.values()]
if len(set(lengths)) == 1:
    print(f"\n✓ همه ستون‌ها {lengths[0]} آیتم دارند - آماده ایجاد DataFrame")
else:
    print("\n✗ خطا: تعداد آیتم‌ها یکسان نیست!")
    print("جزئیات:")
    for key, value in data.items():
        print(f"  {key}: {len(value)}")
    exit(1)

print("="*60)

# ایجاد DataFrame
df = pd.DataFrame(data)

# ذخیره فایل ساده
df.to_excel('WHCP_Parts_List_2_Simple.xlsx', index=False, engine='openpyxl')
print("\n✓ فایل WHCP_Parts_List_2_Simple.xlsx ایجاد شد")

# ایجاد فایل با فرمت‌بندی
wb = Workbook()
ws = wb.active
ws.title = "Parts List"

# اضافه کردن اطلاعات بالای جدول
ws.append(['Required on Site Date by PASARGAD:'])
ws.append([''])
ws.append(['Reminder: Attach to this form all parts lists and drawings as listed in column 9 and 10'])
ws.append([''])
ws.append(['Manufacturers Date:'])
ws.append([''])
ws.append(['Description of Parts to include all parts recommended to'])
ws.append(['be kept for normal operation and slow wearing parts'])
ws.append([''])

# ادغام سلول‌ها
ws.merge_cells('A1:E1')
ws.merge_cells('A3:E3')
ws.merge_cells('A5:E5')
ws.merge_cells('A7:E7')
ws.merge_cells('A8:E8')

# فرمت‌بندی سطرهای بالایی
for row_num in [1, 3, 5, 7, 8]:
    cell = ws[f'A{row_num}']
    if row_num == 1:
        cell.font = Font(bold=True, size=11, color="FF0000")  # قرمز
    else:
        cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# اضافه کردن هدر جدول
current_row = 10
headers = ['Row', 'Qty', 'Description of Parts', 'Material (see Note 4 above)', 
           'Recommended by Manufacturer for 19 MPD']

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# اضافه کردن داده‌ها
for idx, row in df.iterrows():
    current_row += 1
    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = row[col_name]
        cell.border = border
        
        # رنگ‌بندی برای Electrical Component
        if col_num == 4 and 'Electrical Component' in str(row[col_name]):
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# تنظیم عرض ستون‌ها
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 30

# تنظیم ارتفاع سطر هدر
ws.row_dimensions[10].height = 40

# ذخیره فایل
wb.save('WHCP_Parts_List_2_Formatted.xlsx')
print("✓ فایل WHCP_Parts_List_2_Formatted.xlsx ایجاد شد")

print("\n" + "="*60)
print("📊 خلاصه نتایج:")
print("="*60)
print(f"✓ تعداد کل قطعات: {len(df)} قلم")
print(f"✓ مجموع تعداد: {df['Qty'].sum()} عدد")

# نمایش آمار جزئی‌تر
print("\n📈 آمار براساس نوع متریال:")
material_stats = df['Material (see Note 4 above)'].value_counts()
for material, count in material_stats.items():
    qty_sum = df[df['Material (see Note 4 above)'] == material]['Qty'].sum()
    print(f"  • {material}: {count} قلم (مجموع {qty_sum} عدد)")

print("\n" + "="*60)
print("✅ فایل‌ها با موفقیت ایجاد شدند:")
print("   1. WHCP_Parts_List_2_Simple.xlsx")
print("   2. WHCP_Parts_List_2_Formatted.xlsx")
print("="*60)
