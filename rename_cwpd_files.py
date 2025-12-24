#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
اسکریپت تغییر نام فایل‌های Excel CWPD
استخراج اطلاعات از شیت‌های Excel و تولید گزارش
"""

import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class CWPDFileRenamer:
    def __init__(self, source_directory):
        """
        مقداردهی اولیه
        
        Args:
            source_directory: مسیر پوشه حاوی فایل‌های Excel
        """
        self.source_directory = Path(source_directory)
        self.results = []
        
    def extract_data_from_excel(self, file_path):
        """
        استخراج Date از فایل Excel (Number دیگه از فایل نمیخونیم)
        
        Args:
            file_path: مسیر فایل Excel
            
        Returns:
            dict: دیکشنری حاوی date
        """
        data = {
            'date': None,
            'date_obj': None  # برای مرتب‌سازی
        }
        
        try:
            # خواندن با data_only=True
            wb_data = load_workbook(file_path, data_only=True)
            
            # بررسی وجود شیت WPD
            if 'WPD' in wb_data.sheetnames:
                ws_data = wb_data['WPD']
                
                # Date از C5
                date_cell = ws_data['C5']
                if date_cell.value:
                    if isinstance(date_cell.value, datetime):
                        data['date'] = date_cell.value.strftime('%d-%b-%Y')
                        data['date_obj'] = date_cell.value
                    else:
                        data['date'] = str(date_cell.value)
            
            # اگر WPD نبود، از HOME استفاده کن
            elif 'HOME' in wb_data.sheetnames:
                ws_home = wb_data['HOME']
                
                # Date از D10
                date_cell = ws_home['D10']
                if date_cell.value:
                    if isinstance(date_cell.value, datetime):
                        data['date'] = date_cell.value.strftime('%d-%b-%Y')
                        data['date_obj'] = date_cell.value
                    else:
                        data['date'] = str(date_cell.value)
            
            wb_data.close()
            
        except Exception as e:
            print(f"      خطا در خواندن فایل: {str(e)}")
        
        return data
    
    def generate_new_filename(self, number):
        """
        تولید نام جدید فایل
        
        Args:
            number: شماره Ref
            
        Returns:
            str: نام جدید فایل
        """
        # فرمت: SJSC-GGNRSP-EPWC-REDA-[Number]-G00
        if not number:
            number = "0000"
            
        new_name = f"SJSC-GGNRSP-EPWC-REDA-{number}-G00.xlsx"
        return new_name
    
    def process_files(self, rename_files=False):
        """
        پردازش فایل‌های Excel و شماره‌گذاری بر اساس تاریخ
        
        Args:
            rename_files: آیا فایل‌ها تغییر نام پیدا کنند؟
            
        Returns:
            list: لیست نتایج
        """
        # پیدا کردن تمام فایل‌های Excel
        excel_files = list(self.source_directory.glob("*.xlsx"))
        excel_files.extend(list(self.source_directory.glob("*.xls")))
        
        # فیلتر کردن فایل‌های CWPD
        cwpd_files = [f for f in excel_files if f.name.startswith('CWPD-')]
        
        print(f"تعداد {len(cwpd_files)} فایل CWPD پیدا شد.\n")
        
        # مرحله 1: استخراج تاریخ از همه فایل‌ها
        files_with_dates = []
        
        for file_path in cwpd_files:
            if file_path.name.startswith('~$'):
                continue
            
            print(f"📊 در حال خواندن: {file_path.name}")
            
            try:
                data = self.extract_data_from_excel(file_path)
                
                files_with_dates.append({
                    'path': file_path,
                    'date': data['date'],
                    'date_obj': data['date_obj']
                })
                
                print(f"  Date: {data['date']}\n")
                
            except Exception as e:
                print(f"  خطا: {str(e)}\n")
                files_with_dates.append({
                    'path': file_path,
                    'date': None,
                    'date_obj': None
                })
        
        # مرحله 2: مرتب‌سازی بر اساس تاریخ
        print("\n" + "="*70)
        print("مرتب‌سازی فایل‌ها بر اساس تاریخ...")
        print("="*70 + "\n")
        
        # فایل‌هایی که تاریخ دارند
        files_with_valid_dates = [f for f in files_with_dates if f['date_obj'] is not None]
        files_without_dates = [f for f in files_with_dates if f['date_obj'] is None]
        
        # مرتب‌سازی
        files_with_valid_dates.sort(key=lambda x: x['date_obj'])
        
        # ترکیب: ابتدا فایل‌های دارای تاریخ، سپس بدون تاریخ
        sorted_files = files_with_valid_dates + files_without_dates
        
        # مرحله 3: شماره‌گذاری و پردازش
        for index, file_info in enumerate(sorted_files, start=1):
            file_path = file_info['path']
            date_str = file_info['date']
            
            # شماره به ترتیب
            number = str(index).zfill(4)
            
            print(f"📊 [{index}/{len(sorted_files)}] {file_path.name}")
            print(f"  Number: {number} (بر اساس ترتیب تاریخ)")
            print(f"  Date: {date_str}")
            
            # تولید نام جدید
            new_filename = self.generate_new_filename(number)
            
            print(f"  نام جدید: {new_filename}\n")
            
            # ذخیره نتیجه
            result = {
                'original_name': file_path.name,
                'new_name': new_filename,
                'number': number,
                'date': date_str or 'N/A',
                'status': 'پردازش شد'
            }
            
            # تغییر نام فایل (اختیاری)
            if rename_files:
                new_path = file_path.parent / new_filename
                if not new_path.exists():
                    file_path.rename(new_path)
                    result['status'] = 'تغییر نام داده شد'
                else:
                    result['status'] = 'فایل با این نام وجود دارد'
            
            self.results.append(result)
        
        return self.results
    
    def create_excel_report(self, output_path):
        """
        تولید گزارش Excel
        
        Args:
            output_path: مسیر فایل خروجی Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش فایل‌های CWPD"
        
        # تنظیمات سبک
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # هدر جدول
        headers = ['ردیف', 'نام اصلی فایل', 'نام جدید فایل', 'Ref Number', 'Date', 'وضعیت']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # داده‌ها
        for row_idx, result in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1).border = border
            ws.cell(row=row_idx, column=2, value=result['original_name']).border = border
            ws.cell(row=row_idx, column=3, value=result['new_name']).border = border
            ws.cell(row=row_idx, column=4, value=result['number']).border = border
            ws.cell(row=row_idx, column=5, value=result['date']).border = border
            ws.cell(row=row_idx, column=6, value=result['status']).border = border
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        # ذخیره فایل
        wb.save(output_path)
        print(f"\n✓ گزارش Excel در مسیر زیر ذخیره شد:")
        print(f"  {output_path}")


def main():
    """
    تابع اصلی برنامه
    """
    print("=" * 70)
    print("برنامه تغییر نام فایل‌های Excel CWPD")
    print("=" * 70)
    print()
    
    # مسیر پوشه فایل‌ها
    source_dir = r"D:\Sepher_Pasargad\works\Production\01-CWPD"
    
    # بررسی وجود پوشه
    if not os.path.exists(source_dir):
        print(f"خطا: پوشه {source_dir} یافت نشد!")
        print("لطفاً مسیر را در کد بررسی کنید.")
        return
    
    # ایجاد شیء پردازشگر
    renamer = CWPDFileRenamer(source_dir)
    
    # پردازش فایل‌ها (بدون تغییر نام واقعی)
    print("در حال پردازش فایل‌ها...")
    print("-" * 70)
    results = renamer.process_files(rename_files=False)
    
    # تولید گزارش Excel
    output_excel = os.path.join(source_dir, f"گزارش_CWPD_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    renamer.create_excel_report(output_excel)
    
    # خلاصه نتایج
    print("\n" + "=" * 70)
    print("خلاصه نتایج:")
    print(f"  تعداد کل فایل‌ها: {len(results)}")
    print(f"  موفق: {sum(1 for r in results if 'خطا' not in r['status'])}")
    print(f"  خطا: {sum(1 for r in results if 'خطا' in r['status'])}")
    print("=" * 70)
    
    # سوال برای تغییر نام واقعی
    print("\nآیا می‌خواهید فایل‌ها واقعاً تغییر نام پیدا کنند?")
    print("توجه: این عملیات قابل بازگشت نیست!")
    choice = input("برای تغییر نام 'yes' وارد کنید: ")
    
    if choice.lower() == 'yes':
        print("\nدر حال تغییر نام فایل‌ها...")
        renamer.results = []
        results = renamer.process_files(rename_files=True)
        
        # تولید گزارش جدید
        output_excel_final = os.path.join(source_dir, f"گزارش_نهایی_CWPD_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        renamer.create_excel_report(output_excel_final)
        print("\n✓ فایل‌ها با موفقیت تغییر نام یافتند!")


if __name__ == "__main__":
    main()
